from bs4 import BeautifulSoup
import requests
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side


def check_card_prices():
    name = []
    seller = []
    expansion = []
    condition = []
    rarity = []
    language = []
    price = []

    with open('./data/urls.txt') as file:
        for line in file:
            url = line.rstrip()
            if url == "":
                print(url + ' - Invalid URL')
                continue
            html_text = requests.get(str(url)).text
            response = requests.get(str(url))
            if response.status_code == 404:
                print(url + ' - Failed')
                continue
            soup = BeautifulSoup(html_text, 'lxml')
            header = soup.find('div', class_='page-title-container d-flex align-items-center')
            name.append(header.find_next('h1').text)
            # Find the first article on page
            article = soup.find('div', class_='row no-gutters article-row')
            # Add first seller name to list
            seller.append(article.find_next('span', class_='d-flex has-content-centered mr-1').text)
            # Add expansion set name to list
            expansion.append(article.find_next('a', class_='expansion-symbol is-magic icon is-24x24 d-flex mr-1')['title'])
            # Add card condition to list
            condition.append(article.find_next('a', class_=['article-condition condition-mt mr-1',
                                                            'article-condition condition-nm mr-1',
                                                            'article-condition condition-ex mr-1',
                                                            'article-condition condition-gd mr-1',
                                                            'article-condition condition-lp mr-1',
                                                            'article-condition condition-pl mr-1',
                                                            'article-condition condition-po mr-1']).text)
            # Add card rarity to list
            rarity.append(article.find_next('span', class_='icon mr-2')['title'])
            # Create temporary soup find data for language
            temp = article.find('a', class_=['article-condition condition-mt mr-1',
                                             'article-condition condition-nm mr-1', 'article-condition condition-ex mr-1',
                                             'article-condition condition-gd mr-1', 'article-condition condition-lp mr-1',
                                             'article-condition condition-pl mr-1', 'article-condition condition-po mr-1'])
            # Add language information to list
            language.append(temp.find_next('span', class_='icon mr-2')['data-original-title'])
            # Add card price to list
            price.append(article.find_next('span', class_='font-weight-bold color-primary small text-right text-nowrap').text)
            print(url + ' - Done')

    # Pandas dataframe for xlsx storing
    df = pd.DataFrame(
        {'Name:': name, 'Price:': price, 'Condition:': condition, 'Expansion:': expansion, 'Language:': language,
         'Seller:': seller, 'Rarity:': rarity})

    df.sort_values(by=['Name:'], inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Excel port 
    filepath_excel = r"./data/cardprices.xlsx"
    book = load_workbook(filepath_excel)
    date = datetime.now().strftime("%d.%m.%Y - %H%M")
    book.create_sheet(date, 0)
    book.save(filepath_excel)

    writer = pd.ExcelWriter(filepath_excel, engine='openpyxl', mode='a', if_sheet_exists='replace')
    df.to_excel(writer, sheet_name=date)

    # Close the Pandas Excel writer and output the Excel file.
    writer.close()


# Function to format the fetched data of the latest sheet in xlsx
def format_xlsx():
    # Path to xlsx file and open workbook
    filepath_excel = r"./data/cardprices.xlsx"
    wb = load_workbook(filepath_excel)
    ws0 = wb.worksheets[0]

    # Column width setup
    ws0.column_dimensions['A'].width = 5
    ws0.column_dimensions['B'].width = 30
    ws0.column_dimensions['C'].width = 15
    ws0.column_dimensions['D'].width = 15
    ws0.column_dimensions['E'].width = 30
    ws0.column_dimensions['F'].width = 15
    ws0.column_dimensions['G'].width = 20
    ws0.column_dimensions['H'].width = 15

    # Background colors of cells and border style setup
    bg_light = '00ACB9CA'
    bg_dark = '008497B0'
    border_style = Side(border_style='thin', color='00000000')
    border = Border(top=border_style, bottom=border_style, left=border_style, right=border_style)

    # Apply styling to all cells
    for rows in ws0.iter_rows(min_row=1, max_row=ws0.max_row, min_col=1, max_col=8):
        for cell in rows:
            if cell.row % 2:
                cell.fill = PatternFill(start_color=bg_light, end_color=bg_light, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
            else:
                cell.fill = PatternFill(start_color=bg_dark, end_color=bg_dark, fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

    # Save all changes applied to the workbook
    wb.save(filepath_excel)


# Function to compare the latest prices with the ones in the previous fetched data and highlight changes
def compare_price_changes():
    # Path to xlsx file and open workbook
    filepath_excel = r"./data/cardprices.xlsx"
    wb = load_workbook(filepath_excel)
    ws0 = wb.worksheets[0]
    ws1 = wb.worksheets[1]

    # Setup background highlight colors
    bg_green = '0090ee90'
    bg_red = '00ff726f'

    # Check for each row in column B of the latest worksheet whether the value (name) is present in the previous one
    for cell0B in ws0['B']:
        # Skip the first row (identifier)
        if cell0B.value == 'Name:':
            continue
        for cell1B in ws1['B']:
            # Skip the first row (identifier)
            if cell1B.value == 'Name:':
                continue

            # If the value of the latest worksheet is present in the previous one check the values of those rows in
            # column C in both worksheets. Change the background color of that cell in the latest worksheet according to
            # the changes in the values. If the value (price) has gone down mark the cell with green otherwise with red.
            if cell0B.value == cell1B.value:
                cell0C = ws0['C' + str(cell0B.row)]
                cell1C = ws1['C' + str(cell1B.row)]
                if cell0C.value < cell1C.value:
                    cell0C.fill = PatternFill(start_color=bg_green, end_color=bg_green, fill_type="solid")
                elif cell0C.value > cell1C.value:
                    cell0C.fill = PatternFill(start_color=bg_red, end_color=bg_red, fill_type="solid")

    # Save all changes applied to the workbook
    wb.save(filepath_excel)


# Main function to run the card price fetcher
if __name__ == '__main__':
    check_card_prices()
    format_xlsx()
    compare_price_changes()
